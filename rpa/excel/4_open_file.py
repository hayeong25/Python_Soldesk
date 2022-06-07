# 불러올 때는 load_workbook
# 생성 시에는 Workbook
from openpyxl import load_workbook

# 워크북 가져오기
wb = load_workbook("./rpa/excel/sample.xlsx")

# 워크북 활성화
ws = wb.active

# 셀 데이터 출력
for x in range(1, 11):
    for y in range(1, 11):
        print(ws.cell(x, y).value, end=" ")
    print()

# max_row / max_column : 워크시트가 가지고 있는 행, 열 개수 가져오기
for x in range(1, ws.max_row + 1):
    for y in range(1, ws.max_column + 1):
        print(ws.cell(x, y).value, end=" ")
    print()
