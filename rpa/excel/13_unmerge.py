# 셀 병합 취소
from openpyxl import load_workbook

wb = load_workbook("./rpa/excel/merge.xlsx")
ws = wb.active

# 병합 셀 해제
ws.unmerge_cells("B2:D2")

wb.save("./rpa/excel/unmerge.xlsx")
