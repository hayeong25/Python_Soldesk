from openpyxl import Workbook

wb = Workbook()

# 기본 시트 활성화
ws = wb.active

# 시트 이름 변경
ws.title = "test"

# 셀에 데이터 입력
ws["A1"] = 1
ws["A2"] = 2
ws["A3"] = 3
ws["B1"] = 4
ws["B2"] = 5
ws["B3"] = 6

# 셀 값 가져오기
print("A1", ws["A1"])  # A1 셀의 정보가
print("A1", ws["A1"].value)  # 셀 값 가져오려면 .value 반드시

# ws.cell(row, column, value)
print(ws.cell(1, 1).value)
print(ws.cell(1, 2).value)
print(ws.cell(1, 3, 10).value)

wb.save("./rpa/excel/sample.xlsx")
