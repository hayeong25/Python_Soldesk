from openpyxl import Workbook
import random

wb = Workbook
ws = wb.active

# 한 줄씩 데이터 삽입 - 리스트 이용
ws.append(["번호", "영어", "수학"])  # 첫 줄 입력

for i in range(1, 11):
    ws.append([i, random.randint(0, 100), random.randint(0, 100)])

rows = [
    ["이름", "생년월일"],
    ["가나다", "900807"],
    ["곰탱이", "950901"],
    ["찬료리", "950901"],
    ["꼬물이", "950901"],
]

for row in rows:
    ws.append(row)

wb.save("./rpa/excel/range.xlsx")

# 첫 번째 행 가져와서 출력
print(ws["A1"].value, ws["B1"].value, ws["C1"].value)

for j in range(1, 4):
    print(ws.cell(1, j).value, end=" ")

first_row = ws[1]
for cell in first_row:
    print(cell.value, end=" ")

# 2 ~ 6행 가져오기
row_range = ws[2:6]  # 여기서는 마지막 번호 포함
for rows in row_range:
    for cell in rows:
        print(cell.value, end=" ")
    print()

# iter_rows() / iter_cols() : 전체 rows, cols 함수
for row in ws.iter_rows():
    print(row[2].value)
for col in ws.iter_cols():
    print(col[2].value)
