import enum
from openpyxl import Workbook

wb = Workbook()
ws = wb.active

# [학번, 출석, 퀴즈1, 퀴즈2, 중간고사, 기말고사, 프로젝트]
scores = [
    [1, 10, 8, 5, 14, 26, 12],
    [2, 7, 3, 7, 15, 24, 18],
    [3, 9, 5, 8, 8, 12, 4],
    [4, 7, 8, 7, 17, 21, 18],
    [5, 7, 8, 7, 16, 25, 15],
    [6, 3, 5, 8, 8, 17, 0],
    [7, 4, 9, 10, 16, 27, 18],
    [8, 6, 6, 6, 15, 19, 17],
    [9, 10, 10, 9, 19, 30, 19],
]

ws.append(["학번", "출석", "퀴즈1", "퀴즈2", "중간고사", "기말고사", "프로젝트", "총점", "성적"])
for score in scores:
    ws.append(score)

# 퀴즈2 점수 모두 10점으로 수정
for idx, cell in enumerate(ws["D"]):
    if idx == 0:  # 제목 행은 skip해야 하기 때문
        continue
    cell.value = 10

# H열에 총점 (SUM 이용), I열에 성적 정보 추가
# 총점 90점 이상은 A, 80점 이상은 B, 70점 이상은 C, 나머지는 D / 출석 5점 미만은 총점에 관계없이 F
for idx, score in enumerate(scores, start=2):
    ws.cell(row=idx, column=8).value = "=sum(B{0}:G{0})".format(idx)
    total = sum(score[1:]) - score[3] + 10
    grade = None
    if total >= 90:
        grade = "A"
    elif total >= 80:
        grade = "B"
    elif total >= 70:
        grade = "C"
    else:
        grade = "D"

    if score[1] < 5:
        grade = "F"
    ws.cell(row=idx, column=9).value = grade

# for i in range(2, 11):
#     ws["h" + i] = "=sum(b" + i + ":g" + i + ")"
#     if ws["b" + i].value < 5:
#         ws["i" + i] = "F"
#     else:
#         if ws["h" + i].value >= 90:
#             ws["i" + i] = "A"
#         elif ws["h" + i].value >= 80:
#             ws["i" + i] = "B"
#         elif ws["h" + i].value >= 70:
#             ws["i" + i] = "C"
#         else:
#             ws["i" + i] = "D"

wb.save("./rpa/excel/scores.xlsx")
